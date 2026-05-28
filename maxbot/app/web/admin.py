"""Веб-приложение организатора и технического администратора.

Маршруты:
  /admin/login           — общий логин для обеих ролей (роль определяется по email)
  /admin/logout          — выход
  /admin                 — дашборд: список мероприятий
  /admin/events/<id>     — карточка мероприятия: участники, кнопки, экспорт
  /admin/events/<id>/scan — страница QR-скана с камерой браузера
  /admin/scan/verify     — API проверки QR (POST с qr_token)
  /admin/events/<id>/close — закрыть регистрацию
  /admin/events/<id>/notify — отправить push-уведомление участникам
  /admin/events/<id>/export — выгрузка XLSX
  /admin/audit           — журнал аудита (только tech_admin)
"""
from __future__ import annotations

import io
from datetime import datetime, timezone
from functools import wraps
from typing import Optional

from flask import (
    Blueprint,
    abort,
    flash,
    g,
    redirect,
    render_template,
    request,
    send_file,
    session as flask_session,
    url_for,
    jsonify,
)
from sqlalchemy import select

from ..config import Config
from ..db import session_scope
from ..logging_setup import get_logger, mask_id
from ..models import (
    AuditAction,
    AuditLog,
    Notification,
    NotificationType,
    Registration,
    RegistrationStatus,
)
from ..repositories.repos import (
    AdminRepo,
    ControllerRepo,
    EventRepo,
    NotificationRepo,
    OrganizerRepo,
    RegistrationRepo,
    UserRepo,
)
from ..services import audit
from ..services.codes import verify_qr_token
from ..services.max_client import MaxBotClient
from ..services.passwords import hash_password, verify_password

log = get_logger(__name__)

admin_bp = Blueprint(
    "admin",
    __name__,
    url_prefix="/admin",
    template_folder="templates",
    static_folder="static",
)


# ============================================================
# Аутентификация
# ============================================================


def _current_user():
    """Возвращает кортеж (role, id) или None. role ∈ {'organizer', 'admin', 'controller'}."""
    role = flask_session.get("role")
    uid = flask_session.get("uid")
    if not role or not uid:
        return None
    return role, uid


def _can_access_event(session, role: str, uid: int, event_id: int) -> bool:
    """Проверяет, что текущий пользователь имеет право работать с мероприятием.

    - admin: все мероприятия
    - organizer: только свои
    - controller: только привязанные через event_controllers
    """
    ev = EventRepo.get(session, event_id)
    if ev is None:
        return False
    if role == "admin":
        return True
    if role == "organizer":
        return ev.organizer_id == uid
    if role == "controller":
        return ControllerRepo.has_access(session, uid, event_id)
    return False


def login_required(roles: Optional[list[str]] = None):
    def deco(fn):
        @wraps(fn)
        def wrapper(*args, **kwargs):
            cu = _current_user()
            if cu is None:
                return redirect(url_for("admin.login", next=request.path))
            role, _ = cu
            if roles and role not in roles:
                abort(403)
            g.current_role = role
            return fn(*args, **kwargs)

        return wrapper

    return deco


@admin_bp.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "GET":
        return render_template("login.html")

    email = (request.form.get("email") or "").strip().lower()
    password = request.form.get("password") or ""

    if not email or not password:
        flash("Заполните email и пароль", "error")
        return render_template("login.html"), 400

    with session_scope() as s:
        # Пытаемся найти сначала среди админов, потом среди организаторов
        adm = AdminRepo.get_by_email(s, email)
        if adm and adm.is_active and verify_password(password, adm.password_hash):
            flask_session["role"] = "admin"
            flask_session["uid"] = adm.id
            flask_session["name"] = adm.display_name
            audit.write(
                s,
                AuditAction.admin_login,
                admin_id=adm.id,
                ip_address=request.remote_addr,
            )
            log.info("ADMIN login uid=%s", mask_id(adm.id))
            return redirect(request.args.get("next") or url_for("admin.dashboard"))

        org = OrganizerRepo.get_by_email(s, email)
        if org and org.is_active and verify_password(password, org.password_hash):
            flask_session["role"] = "organizer"
            flask_session["uid"] = org.id
            flask_session["name"] = org.full_name
            audit.write(
                s,
                AuditAction.organizer_login,
                organizer_id=org.id,
                ip_address=request.remote_addr,
            )
            log.info("ORGANIZER login uid=%s", mask_id(org.id))
            return redirect(request.args.get("next") or url_for("admin.dashboard"))

        ctrl = ControllerRepo.get_by_email(s, email)
        if ctrl and ctrl.is_active and verify_password(password, ctrl.password_hash):
            flask_session["role"] = "controller"
            flask_session["uid"] = ctrl.id
            flask_session["name"] = ctrl.full_name
            audit.write(
                s,
                AuditAction.controller_login,
                payload={"controller_id": ctrl.id},
                ip_address=request.remote_addr,
            )
            log.info("CONTROLLER login uid=%s", mask_id(ctrl.id))
            return redirect(request.args.get("next") or url_for("admin.dashboard"))

    flash("Неверный email или пароль", "error")
    return render_template("login.html"), 401


@admin_bp.route("/logout")
def logout():
    flask_session.clear()
    return redirect(url_for("admin.login"))


# ============================================================
# Дашборд
# ============================================================


@admin_bp.route("/")
@login_required()
def dashboard():
    role, uid = _current_user()
    with session_scope() as s:
        if role == "admin":
            events = EventRepo.list_all_for_admin(s)
        elif role == "controller":
            events = ControllerRepo.list_events_for_controller(s, uid)
        else:
            events = EventRepo.list_by_organizer(s, uid)
        rows = []
        for ev in events:
            taken = RegistrationRepo.count_confirmed(s, ev.id)
            rows.append(
                {
                    "id": ev.id,
                    "title": ev.title,
                    "date": ev.event_date,
                    "format": ev.format.value,
                    "location": ev.location,
                    "taken": taken,
                    "total": ev.max_participants,
                    "closed": ev.registration_closed,
                }
            )
    return render_template(
        "dashboard.html", events=rows, role=role, name=flask_session.get("name")
    )


# ============================================================
# Карточка мероприятия
# ============================================================


@admin_bp.route("/events/<int:event_id>")
@login_required()
def event_detail(event_id: int):
    role, uid = _current_user()
    with session_scope() as s:
        ev = EventRepo.get(s, event_id)
        if ev is None:
            abort(404)
        if not _can_access_event(s, role, uid, event_id):
            abort(403)

        # Контролёр видит «облегчённую» карточку — только сканер, без списка участников
        if role == "controller":
            event_data = {
                "id": ev.id,
                "title": ev.title,
                "description": ev.description,
                "date": ev.event_date,
                "duration_minutes": ev.duration_minutes,
                "format": ev.format.value,
                "location": ev.location,
                "max": ev.max_participants,
                "taken": RegistrationRepo.count_confirmed(s, event_id),
                "attended": 0,
                "closed": ev.registration_closed,
            }
            return render_template(
                "event_detail.html",
                event=event_data,
                regs=[],
                notifs=[],
                role=role,
                controllers=[],
            )

        regs = RegistrationRepo.list_for_event(s, event_id)
        reg_rows = [
            {
                "id": r.id,
                "code": r.registration_code,
                "username": r.user.username,
                "status": r.status.value,
                "created_at": r.created_at,
                "attended_at": r.attended_at,
                "cancelled_at": r.cancelled_at,
            }
            for r in regs
        ]
        notifs = NotificationRepo.list_for_event(s, event_id)
        notif_rows = [
            {
                "id": n.id,
                "type": n.type.value,
                "body": n.body,
                "sent_at": n.sent_at,
                "delivered": n.delivered_count,
                "failed": n.failed_count,
            }
            for n in notifs
        ]
        controllers = ControllerRepo.list_for_event(s, event_id)
        controller_rows = [
            {"id": c.id, "email": c.email, "full_name": c.full_name, "is_active": c.is_active}
            for c in controllers
        ]
        taken = sum(1 for r in regs if r.status == RegistrationStatus.confirmed)
        attended = sum(1 for r in regs if r.status == RegistrationStatus.attended)

        event_data = {
            "id": ev.id,
            "title": ev.title,
            "description": ev.description,
            "date": ev.event_date,
            "duration_minutes": ev.duration_minutes,
            "format": ev.format.value,
            "location": ev.location,
            "max": ev.max_participants,
            "taken": taken,
            "attended": attended,
            "closed": ev.registration_closed,
        }
    return render_template(
        "event_detail.html",
        event=event_data,
        regs=reg_rows,
        notifs=notif_rows,
        controllers=controller_rows,
        role=role,
    )


@admin_bp.route("/events/<int:event_id>/close", methods=["POST"])
@login_required()
def event_close(event_id: int):
    role, uid = _current_user()
    with session_scope() as s:
        ev = EventRepo.get(s, event_id)
        if ev is None:
            abort(404)
        if role == "organizer" and ev.organizer_id != uid:
            abort(403)
        ev.registration_closed = True
        audit.write(
            s,
            AuditAction.registration_closed,
            organizer_id=uid if role == "organizer" else None,
            admin_id=uid if role == "admin" else None,
            event_id=event_id,
            ip_address=request.remote_addr,
        )
    flash("Регистрация закрыта", "success")
    return redirect(url_for("admin.event_detail", event_id=event_id))


@admin_bp.route("/events/<int:event_id>/open", methods=["POST"])
@login_required()
def event_open(event_id: int):
    """Открыть ранее закрытую регистрацию."""
    role, uid = _current_user()
    with session_scope() as s:
        ev = EventRepo.get(s, event_id)
        if ev is None:
            abort(404)
        if role == "organizer" and ev.organizer_id != uid:
            abort(403)
        EventRepo.open_registration(s, event_id)
        audit.write(
            s,
            AuditAction.event_updated,
            organizer_id=uid if role == "organizer" else None,
            admin_id=uid if role == "admin" else None,
            event_id=event_id,
            payload={"action": "open_registration"},
            ip_address=request.remote_addr,
        )
    flash("Регистрация открыта", "success")
    return redirect(url_for("admin.event_detail", event_id=event_id))


def _parse_event_form() -> Optional[dict]:
    """Разбирает форму создания/редактирования. Возвращает dict или None при ошибке."""
    try:
        title = (request.form.get("title") or "").strip()
        description = (request.form.get("description") or "").strip()
        date_str = (request.form.get("event_date") or "").strip()
        duration = int(request.form.get("duration_minutes") or "0")
        format_val = (request.form.get("format") or "offline").strip()
        location = (request.form.get("location") or "").strip()
        max_p = int(request.form.get("max_participants") or "0")

        if not title or not description or not date_str or not location:
            flash("Заполните все обязательные поля", "error")
            return None
        if duration <= 0:
            flash("Длительность должна быть положительной", "error")
            return None
        if max_p <= 0:
            flash("Лимит участников должен быть положительным", "error")
            return None
        try:
            from ..models import EventFormat as _EF  # локально, чтобы не циклить
            fmt = _EF(format_val)
        except ValueError:
            flash("Неверный формат мероприятия", "error")
            return None

        # datetime-local приходит как "2026-06-01T11:00"
        try:
            naive = datetime.strptime(date_str, "%Y-%m-%dT%H:%M")
        except ValueError:
            flash("Неверный формат даты", "error")
            return None
        # Считаем что время указано в локальной зоне сервера; делаем aware UTC.
        # Для MVP в большинстве случаев сервер в Moscow — этого достаточно.
        event_dt = naive.replace(tzinfo=timezone.utc)

        return {
            "title": title,
            "description": description,
            "event_date": event_dt,
            "duration_minutes": duration,
            "format_": fmt,
            "location": location,
            "max_participants": max_p,
        }
    except Exception as e:
        flash(f"Ошибка обработки формы: {e}", "error")
        return None


@admin_bp.route("/events/new", methods=["GET", "POST"])
@login_required()
def event_new():
    role, uid = _current_user()
    if request.method == "GET":
        return render_template("event_form.html", mode="new", event=None)

    data = _parse_event_form()
    if data is None:
        return render_template("event_form.html", mode="new", event=None), 400

    with session_scope() as s:
        if role == "organizer":
            organizer_id = uid
        else:
            # Админ создаёт мероприятие от имени первого активного организатора.
            # В будущем можно добавить выбор организатора в форме.
            from ..models import Organizer as _Org
            first_org = s.scalar(select(_Org).where(_Org.is_active.is_(True)).limit(1))
            if first_org is None:
                flash(
                    "Нет ни одного активного организатора. Создайте его через "
                    "python -m migrations.create_users",
                    "error",
                )
                return redirect(url_for("admin.dashboard"))
            organizer_id = first_org.id

        ev = EventRepo.create(s, organizer_id=organizer_id, **data)
        audit.write(
            s,
            AuditAction.event_created,
            organizer_id=uid if role == "organizer" else None,
            admin_id=uid if role == "admin" else None,
            event_id=ev.id,
            payload={"title": data["title"]},
            ip_address=request.remote_addr,
        )
        flash(f"Мероприятие «{data['title']}» создано", "success")
        return redirect(url_for("admin.event_detail", event_id=ev.id))


@admin_bp.route("/events/<int:event_id>/edit", methods=["GET", "POST"])
@login_required()
def event_edit(event_id: int):
    role, uid = _current_user()
    with session_scope() as s:
        ev = EventRepo.get(s, event_id)
        if ev is None:
            abort(404)
        if role == "organizer" and ev.organizer_id != uid:
            abort(403)

        if request.method == "GET":
            event_data = {
                "id": ev.id,
                "title": ev.title,
                "description": ev.description,
                "event_date": ev.event_date,
                "duration_minutes": ev.duration_minutes,
                "format": ev.format.value,
                "location": ev.location,
                "max_participants": ev.max_participants,
            }
            return render_template("event_form.html", mode="edit", event=event_data)

    data = _parse_event_form()
    if data is None:
        # Перерисуем форму с введёнными данными — простой вариант: вернуть с GET
        return redirect(url_for("admin.event_edit", event_id=event_id))

    with session_scope() as s:
        ev = EventRepo.get(s, event_id)
        if ev is None:
            abort(404)
        if role == "organizer" and ev.organizer_id != uid:
            abort(403)
        EventRepo.update(s, event_id, **data)
        audit.write(
            s,
            AuditAction.event_updated,
            organizer_id=uid if role == "organizer" else None,
            admin_id=uid if role == "admin" else None,
            event_id=event_id,
            payload={"title": data["title"]},
            ip_address=request.remote_addr,
        )
        flash("Изменения сохранены", "success")
        return redirect(url_for("admin.event_detail", event_id=event_id))


@admin_bp.route("/events/<int:event_id>/delete", methods=["POST"])
@login_required()
def event_delete(event_id: int):
    """Удаляет мероприятие если на нём нет участников. Если есть — не даёт."""
    role, uid = _current_user()
    with session_scope() as s:
        ev = EventRepo.get(s, event_id)
        if ev is None:
            abort(404)
        if role == "organizer" and ev.organizer_id != uid:
            abort(403)
        regs_count = RegistrationRepo.count_confirmed(s, event_id)
        if regs_count > 0:
            flash(
                f"Нельзя удалить: уже записано {regs_count} участников. "
                "Сначала закройте регистрацию или отмените записи.",
                "error",
            )
            return redirect(url_for("admin.event_detail", event_id=event_id))
        EventRepo.delete(s, event_id)
        audit.write(
            s,
            AuditAction.event_updated,
            organizer_id=uid if role == "organizer" else None,
            admin_id=uid if role == "admin" else None,
            event_id=event_id,
            payload={"action": "deleted"},
            ip_address=request.remote_addr,
        )
    flash("Мероприятие удалено", "success")
    return redirect(url_for("admin.dashboard"))


@admin_bp.route("/events/<int:event_id>/notify", methods=["POST"])
@login_required(roles=["organizer", "admin"])
def event_notify(event_id: int):
    role, uid = _current_user()
    body_text = (request.form.get("body") or "").strip()
    type_val = request.form.get("type") or "info"
    if not body_text:
        flash("Пустое сообщение", "error")
        return redirect(url_for("admin.event_detail", event_id=event_id))

    try:
        ntype = NotificationType(type_val)
    except ValueError:
        flash("Некорректный тип уведомления", "error")
        return redirect(url_for("admin.event_detail", event_id=event_id))

    with session_scope() as s:
        ev = EventRepo.get(s, event_id)
        if ev is None:
            abort(404)
        if role == "organizer" and ev.organizer_id != uid:
            abort(403)

        regs = [
            r
            for r in RegistrationRepo.list_for_event(s, event_id)
            if r.status == RegistrationStatus.confirmed and r.notifications_enabled
        ]

        n = NotificationRepo.create(
            s,
            event_id=event_id,
            organizer_id=ev.organizer_id,
            type_=ntype,
            body=body_text,
        )
        client = MaxBotClient()
        delivered = 0
        failed = 0
        text = f"🔔 «{ev.title}»\n\n{body_text}"
        for r in regs:
            try:
                client.send_message(r.user.max_user_id, text)
                delivered += 1
            except Exception as e:
                failed += 1
                log.warning("notify fail user=%s: %s", mask_id(r.user.max_user_id), e)
        NotificationRepo.update_stats(s, n.id, delivered=delivered, failed=failed)

        audit.write(
            s,
            AuditAction.notification_sent,
            organizer_id=uid if role == "organizer" else None,
            admin_id=uid if role == "admin" else None,
            event_id=event_id,
            payload={
                "type": ntype.value,
                "delivered": delivered,
                "failed": failed,
            },
            ip_address=request.remote_addr,
        )

    flash(f"Отправлено: {delivered}, ошибок: {failed}", "success")
    return redirect(url_for("admin.event_detail", event_id=event_id))


# ============================================================
# QR-сканер
# ============================================================


@admin_bp.route("/events/<int:event_id>/scan")
@login_required()
def event_scan(event_id: int):
    role, uid = _current_user()
    with session_scope() as s:
        ev = EventRepo.get(s, event_id)
        if ev is None:
            abort(404)
        if not _can_access_event(s, role, uid, event_id):
            abort(403)
        event_data = {"id": ev.id, "title": ev.title}
    return render_template("scan.html", event=event_data)


@admin_bp.route("/scan/verify", methods=["POST"])
@login_required()
def scan_verify():
    """API проверки QR. Принимает {token} в JSON, возвращает {ok, reason, participant}."""
    role, uid = _current_user()
    data = request.get_json(silent=True) or {}
    token = (data.get("token") or "").strip()
    event_id = data.get("event_id")

    if not token or not event_id:
        return jsonify({"ok": False, "reason": "Нет токена"}), 400

    reg_id = verify_qr_token(token)
    if reg_id is None:
        log.warning("SCAN bad signature event=%s", event_id)
        return jsonify({"ok": False, "reason": "Подпись QR не прошла"})

    with session_scope() as s:
        reg = RegistrationRepo.get(s, reg_id)
        if reg is None:
            return jsonify({"ok": False, "reason": "Запись не найдена"})
        if reg.event_id != int(event_id):
            return jsonify({"ok": False, "reason": "QR от другого мероприятия"})
        if not _can_access_event(s, role, uid, int(event_id)):
            return jsonify({"ok": False, "reason": "Нет прав на мероприятие"}), 403
        if reg.status == RegistrationStatus.cancelled:
            return jsonify({"ok": False, "reason": "Запись отменена"})
        if reg.status == RegistrationStatus.attended:
            return jsonify(
                {
                    "ok": False,
                    "reason": "Уже отмечен пришедшим",
                    "participant": reg.user.username,
                    "code": reg.registration_code,
                }
            )

        ok = RegistrationRepo.mark_attended(s, reg.id)
        if not ok:
            return jsonify({"ok": False, "reason": "Не удалось отметить"})

        audit.write(
            s,
            AuditAction.attended_marked,
            organizer_id=uid if role == "organizer" else None,
            admin_id=uid if role == "admin" else None,
            event_id=reg.event_id,
            registration_id=reg.id,
            payload={"controller_id": uid} if role == "controller" else None,
            ip_address=request.remote_addr,
        )
        return jsonify(
            {
                "ok": True,
                "reason": "Допущен",
                "participant": reg.user.username,
                "code": reg.registration_code,
            }
        )


# ============================================================
# Экспорт
# ============================================================


@admin_bp.route("/events/<int:event_id>/export")
@login_required()
def event_export(event_id: int):
    from openpyxl import Workbook  # импорт здесь, чтобы не грузить при импорте модуля

    role, uid = _current_user()
    with session_scope() as s:
        ev = EventRepo.get(s, event_id)
        if ev is None:
            abort(404)
        if role == "organizer" and ev.organizer_id != uid:
            abort(403)

        regs = RegistrationRepo.list_for_event(s, event_id)
        wb = Workbook()
        ws = wb.active
        ws.title = "Участники"
        ws.append(["Код записи", "Имя профиля MAX", "Статус", "Записан", "Пришёл"])
        for r in regs:
            ws.append(
                [
                    r.registration_code,
                    r.user.username,
                    r.status.value,
                    r.created_at.strftime("%Y-%m-%d %H:%M"),
                    r.attended_at.strftime("%Y-%m-%d %H:%M") if r.attended_at else "",
                ]
            )

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

    filename = f"event_{event_id}_participants.xlsx"
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ============================================================
# Журнал аудита (только tech admin)
# ============================================================


@admin_bp.route("/audit")
@login_required(roles=["admin"])
def audit_log():
    with session_scope() as s:
        stmt = (
            select(AuditLog)
            .order_by(AuditLog.created_at.desc())
            .limit(200)
        )
        entries = list(s.scalars(stmt))
        rows = [
            {
                "id": e.id,
                "created_at": e.created_at,
                "action": e.action.value,
                "actor_user_id": e.actor_user_id,
                "actor_organizer_id": e.actor_organizer_id,
                "actor_admin_id": e.actor_admin_id,
                "target_event_id": e.target_event_id,
                "target_registration_id": e.target_registration_id,
                "payload": e.payload,
                "ip_address": e.ip_address,
            }
            for e in entries
        ]
    return render_template("audit.html", rows=rows)


# ============================================================
# Управление контролёрами мероприятия (organizer/admin)
# ============================================================


@admin_bp.route("/events/<int:event_id>/controllers/upload", methods=["POST"])
@login_required(roles=["organizer", "admin"])
def event_controllers_upload(event_id: int):
    """Загрузка списка контролёров из XLSX-файла.

    Доступно только организатору мероприятия. Техадмин в этот процесс
    не вмешивается — он только создаёт учётки контролёров в /admin/users.

    Формат файла: первая строка — заголовки (игнорируются),
    далее по строке на контролёра: A=email, B=full_name.
    Если контролёр с таким email существует — привязываем к мероприятию.
    Если нет — добавляем в список ошибок, ничего не создаём.
    """
    from openpyxl import load_workbook

    role, uid = _current_user()
    with session_scope() as s:
        ev = EventRepo.get(s, event_id)
        if ev is None:
            abort(404)
        if role == "organizer" and ev.organizer_id != uid:
            abort(403)

        file = request.files.get("file")
        if file is None or not file.filename:
            flash("Файл не выбран", "error")
            return redirect(url_for("admin.event_detail", event_id=event_id))

        try:
            wb = load_workbook(file, read_only=True, data_only=True)
        except Exception as e:
            flash(f"Не удалось открыть XLSX: {e}", "error")
            return redirect(url_for("admin.event_detail", event_id=event_id))

        ws = wb.active
        if ws is None:
            flash("В XLSX нет листов", "error")
            return redirect(url_for("admin.event_detail", event_id=event_id))

        assigned = []
        not_found = []
        already = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue  # шапка
            if not row or not row[0]:
                continue
            email_raw = str(row[0]).strip().lower()
            # full_name из файла — нам не нужен для поиска, контролёр уже создан техадмином
            ctrl = ControllerRepo.get_by_email(s, email_raw)
            if ctrl is None or not ctrl.is_active:
                not_found.append(email_raw)
                continue
            created = ControllerRepo.assign_to_event(
                s, event_id=event_id, controller_id=ctrl.id,
                granted_by_organizer_id=uid if role == "organizer" else None,
            )
            if created:
                assigned.append(ctrl.full_name)
            else:
                already.append(ctrl.full_name)

        if assigned:
            audit.write(
                s,
                AuditAction.controllers_assigned,
                organizer_id=uid if role == "organizer" else None,
                admin_id=uid if role == "admin" else None,
                event_id=event_id,
                payload={"count": len(assigned)},
                ip_address=request.remote_addr,
            )

    parts = []
    if assigned:
        parts.append(f"✅ Привязано: {len(assigned)} ({', '.join(assigned)})")
    if already:
        parts.append(f"ℹ️ Уже были привязаны: {len(already)}")
    if not_found:
        parts.append(
            f"❌ Не найдено: {len(not_found)} ({', '.join(not_found)}). "
            "Обратитесь к техадмину для создания учёток."
        )
    flash(" · ".join(parts) if parts else "Файл пуст", "success" if assigned else "error")
    return redirect(url_for("admin.event_detail", event_id=event_id))


@admin_bp.route("/events/<int:event_id>/controllers/<int:controller_id>/revoke", methods=["POST"])
@login_required(roles=["organizer", "admin"])
def event_controllers_revoke(event_id: int, controller_id: int):
    role, uid = _current_user()
    with session_scope() as s:
        ev = EventRepo.get(s, event_id)
        if ev is None:
            abort(404)
        if role == "organizer" and ev.organizer_id != uid:
            abort(403)
        ok = ControllerRepo.revoke_from_event(s, event_id, controller_id)
        if ok:
            audit.write(
                s,
                AuditAction.controller_revoked,
                organizer_id=uid if role == "organizer" else None,
                admin_id=uid if role == "admin" else None,
                event_id=event_id,
                payload={"controller_id": controller_id},
                ip_address=request.remote_addr,
            )
            flash("Доступ отозван", "success")
        else:
            flash("Контролёр не был привязан", "error")
    return redirect(url_for("admin.event_detail", event_id=event_id))


# ============================================================
# Управление пользователями (только tech admin)
# ============================================================


@admin_bp.route("/users", methods=["GET", "POST"])
@login_required(roles=["admin"])
def users_management():
    role, uid = _current_user()

    if request.method == "POST":
        email = (request.form.get("email") or "").strip().lower()
        full_name = (request.form.get("full_name") or "").strip()
        password = request.form.get("password") or ""
        role_val = (request.form.get("role") or "").strip()

        if not email or not full_name or not password or not role_val:
            flash("Заполните все поля", "error")
            return redirect(url_for("admin.users_management"))

        if role_val not in ("organizer", "controller", "admin"):
            flash("Неверная роль", "error")
            return redirect(url_for("admin.users_management"))

        if len(password) < 6:
            flash("Пароль должен быть не короче 6 символов", "error")
            return redirect(url_for("admin.users_management"))

        with session_scope() as s:
            # Проверка что email не занят ни в одной из таблиц
            if (OrganizerRepo.get_by_email(s, email)
                    or ControllerRepo.get_by_email(s, email)
                    or AdminRepo.get_by_email(s, email)):
                flash(f"Email {email} уже используется", "error")
                return redirect(url_for("admin.users_management"))

            ph = hash_password(password)
            if role_val == "organizer":
                obj = OrganizerRepo.create(s, email=email, password_hash=ph, full_name=full_name)
                payload = {"role": "organizer", "organizer_id": obj.id}
            elif role_val == "controller":
                obj = ControllerRepo.create(s, email=email, password_hash=ph, full_name=full_name)
                payload = {"role": "controller", "controller_id": obj.id}
            else:  # admin
                obj = AdminRepo.create(s, email=email, password_hash=ph, display_name=full_name)
                payload = {"role": "admin", "admin_id": obj.id}

            audit.write(
                s,
                AuditAction.user_created,
                admin_id=uid,
                payload=payload,
                ip_address=request.remote_addr,
            )

        flash(f"Создан {role_val}: {email}", "success")
        return redirect(url_for("admin.users_management"))

    # GET — список всех учёток
    with session_scope() as s:
        organizers = OrganizerRepo.list_all(s)
        controllers = ControllerRepo.list_all(s)
        admins = AdminRepo.list_all(s)

        org_rows = [
            {"id": o.id, "email": o.email, "full_name": o.full_name, "is_active": o.is_active}
            for o in organizers
        ]
        ctrl_rows = [
            {"id": c.id, "email": c.email, "full_name": c.full_name, "is_active": c.is_active}
            for c in controllers
        ]
        adm_rows = [
            {"id": a.id, "email": a.email, "display_name": a.display_name, "is_active": a.is_active}
            for a in admins
        ]

    return render_template(
        "users.html",
        organizers=org_rows,
        controllers=ctrl_rows,
        admins=adm_rows,
    )


@admin_bp.route("/users/<role_val>/<int:user_id>/edit", methods=["POST"])
@login_required(roles=["admin"])
def user_edit(role_val: str, user_id: int):
    """Редактирование учётки: меняем ФИО и/или пароль.
    Email и роль не меняются (см. ТЗ).
    """
    role, uid = _current_user()
    if role_val not in ("organizer", "controller", "admin"):
        abort(404)

    full_name = (request.form.get("full_name") or "").strip()
    new_password = request.form.get("new_password") or ""

    if not full_name:
        flash("ФИО не может быть пустым", "error")
        return redirect(url_for("admin.users_management"))

    if new_password and len(new_password) < 6:
        flash("Пароль должен быть не короче 6 символов", "error")
        return redirect(url_for("admin.users_management"))

    with session_scope() as s:
        if role_val == "organizer":
            obj = OrganizerRepo.get(s, user_id)
            if obj is None:
                abort(404)
            obj.full_name = full_name
            target_email = obj.email
        elif role_val == "controller":
            obj = ControllerRepo.get(s, user_id)
            if obj is None:
                abort(404)
            obj.full_name = full_name
            target_email = obj.email
        else:  # admin
            obj = AdminRepo.get(s, user_id)
            if obj is None:
                abort(404)
            obj.display_name = full_name
            target_email = obj.email

        changes = {"full_name_changed": True}
        if new_password:
            obj.password_hash = hash_password(new_password)
            changes["password_changed"] = True

        audit.write(
            s,
            AuditAction.admin_action,
            admin_id=uid,
            payload={
                "operation": "user_edit",
                "target_role": role_val,
                "target_id": user_id,
                **changes,
            },
            ip_address=request.remote_addr,
        )

    flash(f"Учётка {target_email} обновлена", "success")
    return redirect(url_for("admin.users_management"))


@admin_bp.route("/users/<role_val>/<int:user_id>/toggle", methods=["POST"])
@login_required(roles=["admin"])
def user_toggle(role_val: str, user_id: int):
    """Переключение активности: включить/отключить учётку.

    Это «мягкое удаление». При деактивации:
    - пользователь не сможет войти в админку
    - запись в БД остаётся (включая историю в audit_log)
    - можно снова активировать обратно
    """
    role, uid = _current_user()
    if role_val not in ("organizer", "controller", "admin"):
        abort(404)

    # Защита от само-деактивации админа
    if role_val == "admin" and user_id == uid:
        flash("Нельзя отключить свою же учётку. Зайдите под другим админом.", "error")
        return redirect(url_for("admin.users_management"))

    with session_scope() as s:
        if role_val == "organizer":
            obj = OrganizerRepo.get(s, user_id)
        elif role_val == "controller":
            obj = ControllerRepo.get(s, user_id)
        else:
            obj = AdminRepo.get(s, user_id)
        if obj is None:
            abort(404)

        new_state = not obj.is_active
        obj.is_active = new_state

        audit.write(
            s,
            AuditAction.user_deactivated if not new_state else AuditAction.admin_action,
            admin_id=uid,
            payload={
                "operation": "user_activate" if new_state else "user_deactivate",
                "target_role": role_val,
                "target_id": user_id,
                "new_state": "active" if new_state else "inactive",
            },
            ip_address=request.remote_addr,
        )

        target_email = obj.email

    flash(
        f"Учётка {target_email} {'активирована' if new_state else 'отключена'}",
        "success",
    )
    return redirect(url_for("admin.users_management"))
